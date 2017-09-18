# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.http import HttpResponse,HttpResponseRedirect
from django.template import RequestContext
from django.db.models.query_utils import Q
from django.contrib.auth.decorators import login_required
from libs.sendmail import send_mail
from KPI.models import table,table_detail,ban
from vacation.models import user_table, state
from wzhl_oa.settings import BASE_DIR
from openpyxl import load_workbook
import simplejson
import xlsxwriter
import datetime
import os
from threading import Thread

import sys
reload(sys)
sys.setdefaultencoding('utf-8')

@login_required
def KPI_table(request):
    path = request.path.split('/')[1]

    orm = ban.objects.all().order_by('id').reverse()
    if orm:
        expire_info = '''{0} 的截止日期为 {1}'''.format(orm[0].KPI_name,str(orm[0].ban_date).split('+')[0])
    else:
        expire_info = ''

    return render(request, 'KPI/KPI_table.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'KPI',
                                                 'path2':path,
                                                 'page_name1':u'绩效管理',
                                                 'page_name2':u'绩效考评',
                                                 'username':request.user.username,
                                                 'expire_info':expire_info},
                                                context_instance=RequestContext(request))

@login_required
def KPI_table_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['KPI_name','final_score','KPI_level','status_interface',None,'name','status','id']

    if  sSortDir_0 == 'asc':
        if sSearch == '':
            result_data = table.objects.filter(name=request.user.first_name).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).count()
        else:
            result_data = table.objects.filter(name=request.user.first_name).filter(Q(KPI_name__contains=sSearch) | \
                                                                                    Q(final_score__contains=sSearch) | \
                                                                                    Q(KPI_level__contains=sSearch) | \
                                                                                    Q(status_interface__contains=sSearch)) \
                                                                                .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).filter(Q(KPI_name__contains=sSearch) | \
                                                                                    Q(final_score__contains=sSearch) | \
                                                                                    Q(KPI_level__contains=sSearch) | \
                                                                                    Q(status_interface__contains=sSearch)).count()
    else:
        if sSearch == '':
            result_data = table.objects.filter(name=request.user.first_name).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).count()
        else:
            result_data = table.objects.filter(name=request.user.first_name).filter(Q(KPI_name__contains=sSearch) | \
                                                                                    Q(final_score__contains=sSearch) | \
                                                                                    Q(KPI_level__contains=sSearch) | \
                                                                                    Q(status_interface__contains=sSearch)) \
                                                    .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).filter(Q(KPI_name__contains=sSearch) | \
                                                                                    Q(final_score__contains=sSearch) | \
                                                                                    Q(KPI_level__contains=sSearch) | \
                                                                                    Q(status_interface__contains=sSearch)).count()



    for i in  result_data:
        export = '''
                <a class="btn btn-sm green">
                    生成Excel文件 <i class="fa fa-level-down"></i>
                </a>
            '''.format(i.id)
        aaData.append({
                       '0':i.KPI_name,
                       '1':i.final_score,
                       '2':i.KPI_level,
                       '3':i.status_interface,
                       '4':export,
                       '5':i.name,
                       '6':i.status,
                       '7':i.id
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(simplejson.dumps(result),content_type="application/json")

@login_required
def KPI_set_session(request):
    KPI_name = request.POST.get('KPI_name')
    name = request.POST.get('name')
    status = request.POST.get('status')
    request.session['KPI'] = KPI_name,name
    if int(status) == 1:
        request.session['status'] = 1
    if int(status) == 2:
        request.session['status'] = 2
    if int(status) == 3:
        request.session['status'] = 3
    if int(status) == 4:
        request.session['status'] = 4
    if int(status) == 5:
        request.session['status'] = 5
    if int(status) == 6:
        request.session['status'] = 6
    if int(status) == 7:
        request.session['status'] = 7
    return HttpResponse(simplejson.dumps('OK'),content_type="application/json")

@login_required
def KPI_table_export_all(request):
    try:
        year = request.POST.get('year')
        workbook = xlsxwriter.Workbook(BASE_DIR + '/static/files/KPI_year.xlsx'.format(year))
        worksheet1 = workbook.add_worksheet()
        # worksheet2 = workbook.add_worksheet('运营')
        sheet1_count = 3
        # sheet2_count = 3


        title = [u'姓名',u'部门',u'上级主管',u'部门负责人',u'入职时间',u'{0}Q1分数'.format(year),
                 u'{0}Q1等级'.format(year),u'{0}Q2分数'.format(year),u'{0}Q2等级'.format(year),
                 u'{0}Q3分数'.format(year),u'{0}Q3等级'.format(year),u'{0}Q4分数'.format(year),
                 u'{0}Q4等级'.format(year)]
        worksheet1.write_row('B2',title)

        vacation_user_table_orm = user_table.objects.all()
        for i in vacation_user_table_orm:

            KPI_table_orm = table.objects.filter(name=i.name)
            # flag = 0
            row = [i.name,i.department,i.supervisor,i.principal,str(i.join_date).split('+')[0],'','','','','','','','']
            for j in KPI_table_orm:
                if j.KPI_name[0:4] == year:
                    if j.KPI_name[-1] == '1':
                        row[5] = j.final_score
                        row[6] = j.KPI_level
                    if j.KPI_name[-1] == '2':
                        row[7] = j.final_score
                        row[8] = j.KPI_level
                    if j.KPI_name[-1] == '3':
                        row[9] = j.final_score
                        row[10] = j.KPI_level
                    if j.KPI_name[-1] == '4':
                        row[11] = j.final_score
                        row[12] = j.KPI_level
                # else:
                #     flag = 2
                #     title = [u'姓名',u'部门',u'上级主管',u'部门负责人',u'入职时间',u'{0}M1'.format(year),
                #              u'{0}M2'.format(year),u'{0}M3'.format(year),u'{0}M4'.format(year),u'{0}M5'.format(year),
                #              u'{0}M6'.format(year),u'{0}M7'.format(year),u'{0}M8'.format(year),u'{0}M9'.format(year),
                #              u'{0}M10'.format(year),u'{0}M11'.format(year),u'{0}M12'.format(year)]
                #     worksheet2.write_row('B2',title)
                #     row = [i.name,i.department,i.supervisor,i.principal,str(i.join_date).split('+')[0]]
                #     if j.KPI_name[0:4] == year:
                #         month = j.KPI_name.split('M')[1]
                #         if month == '1':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '2':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '3':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '4':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '5':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '6':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '7':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '8':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '9':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '10':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '11':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
                #         if month == '12':
                #             row.append(j.final_score)
                #         else:
                #             row.append('')
            # if flag == 1:
            worksheet1.write_row('B%s' % sheet1_count, row)
            sheet1_count += 1
            # elif flag == 2:
            #     worksheet2.write_row('B%s' % sheet2_count, row)
            #     sheet2_count += 1
            # else:
            #     pass
        workbook.close()
        return HttpResponse(simplejson.dumps({'code':0,'msg':u'生成Excel文件成功'}),content_type="application/json")
    except Exception,e:
        print e
        return HttpResponse(simplejson.dumps({'code':1,'msg':u'生成Excel文件失败'}),content_type="application/json")

@login_required
def KPI_table_detail(request):
    if request.session.get('KPI'):
        KPI_name = request.session.get('KPI')[0]
        name = request.session.get('KPI')[1]
    else:
        KPI_name = ''
        name = ''
    orm = table.objects.filter(KPI_name=KPI_name).filter(name=name)

    if len(orm):
        for i in orm:
            self_comment = i.self_comment.replace('\n','\\n')
            supervisor_comment = i.supervisor_comment.replace('\n','\\n')
            principal_comment = i.principal_comment.replace('\n','\\n')
    else:
        self_comment = ''
        supervisor_comment = ''
        principal_comment = ''
    # try:
    #     if status:pass
    # except Exception:
    #     status = '员工设定目标'
    return render(request, 'KPI/KPI_table_detail.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'KPI',
                                                 'path2':'',
                                                 'page_name1':u'绩效管理',
                                                 'page_name2':u'绩效考评详情',
                                                 'self_comment':self_comment,
                                                 'supervisor_comment':supervisor_comment,
                                                 'principal_comment':principal_comment,
                                                 'KPI_name':KPI_name,
                                                 'name':name,
                                                 'self_comment_row':len(self_comment.split('\\n')) + 2,
                                                 'supervisor_comment_row':len(supervisor_comment.split('\\n')) + 2,
                                                 'principal_comment_row':len(principal_comment.split('\\n')) + 2},
                                                context_instance=RequestContext(request))

@login_required
def KPI_table_detail_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    if request.session.get('KPI'):
        KPI_name = request.session.get('KPI')[0]
        name = request.session.get('KPI')[1]
    else:
        result = {'sEcho':sEcho,
               'iTotalRecords':0,
               'iTotalDisplayRecords':0,
               'aaData':[]
        }
        return HttpResponse(simplejson.dumps(result),content_type="application/json")

    aaData = []
    sort = ['KPI_name','final_score','objective','description','weight','self_report_value','self_report_score',
            'supervisor_report_value','supervisor_report_score','principal_report_value','principal_report_score','id']

    if  sSortDir_0 == 'asc':
        if sSearch == '':
            result_data = table_detail.objects.filter(KPI_name=KPI_name).filter(name=name).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table_detail.objects.filter(KPI_name=KPI_name).filter(name=name).count()
        else:
            result_data = table_detail.objects.filter(KPI_name=KPI_name).filter(name=name).filter(Q(KPI_name__contains=sSearch) | \
                                                    Q(final_score__contains=sSearch) | \
                                                    Q(objective__contains=sSearch) | \
                                                    Q(description__contains=sSearch) | \
                                                    Q(weight__contains=sSearch) | \
                                                    Q(self_report_value__contains=sSearch) | \
                                                    Q(self_report_score__contains=sSearch) | \
                                                    Q(supervisor_report_value__contains=sSearch) | \
                                                    Q(supervisor_report_score__contains=sSearch) | \
                                                    Q(principal_report_value__contains=sSearch) | \
                                                    Q(principal_report_score__contains=sSearch)) \
                                                    .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table_detail.objects.filter(KPI_name=KPI_name).filter(name=name).filter(Q(KPI_name__contains=sSearch) | \
                                                    Q(final_score__contains=sSearch) | \
                                                    Q(objective__contains=sSearch) | \
                                                    Q(description__contains=sSearch) | \
                                                    Q(weight__contains=sSearch) | \
                                                    Q(self_report_value__contains=sSearch) | \
                                                    Q(self_report_score__contains=sSearch) | \
                                                    Q(supervisor_report_value__contains=sSearch) | \
                                                    Q(supervisor_report_score__contains=sSearch) | \
                                                    Q(principal_report_value__contains=sSearch) | \
                                                    Q(principal_report_score__contains=sSearch)).count()
    else:
        if sSearch == '':
            result_data = table_detail.objects.filter(KPI_name=KPI_name).filter(name=name).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table_detail.objects.filter(KPI_name=KPI_name).filter(name=name).count()
        else:
            result_data = table_detail.objects.filter(KPI_name=KPI_name).filter(name=name).filter(Q(KPI_name__contains=sSearch) | \
                                                    Q(final_score__contains=sSearch) | \
                                                    Q(objective__contains=sSearch) | \
                                                    Q(description__contains=sSearch) | \
                                                    Q(weight__contains=sSearch) | \
                                                    Q(self_report_value__contains=sSearch) | \
                                                    Q(self_report_score__contains=sSearch) | \
                                                    Q(supervisor_report_value__contains=sSearch) | \
                                                    Q(supervisor_report_score__contains=sSearch) | \
                                                    Q(principal_report_value__contains=sSearch) | \
                                                    Q(principal_report_score__contains=sSearch)) \
                                                    .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table_detail.objects.filter(KPI_name=KPI_name).filter(name=name).filter(Q(KPI_name__contains=sSearch) | \
                                                    Q(final_score__contains=sSearch) | \
                                                    Q(objective__contains=sSearch) | \
                                                    Q(description__contains=sSearch) | \
                                                    Q(weight__contains=sSearch) | \
                                                    Q(self_report_value__contains=sSearch) | \
                                                    Q(self_report_score__contains=sSearch) | \
                                                    Q(supervisor_report_value__contains=sSearch) | \
                                                    Q(supervisor_report_score__contains=sSearch) | \
                                                    Q(principal_report_value__contains=sSearch) | \
                                                    Q(principal_report_score__contains=sSearch)).count()

    for i in  result_data:
        objective = '<br>'.join(i.objective.split('\n')).replace(' ','&nbsp')
        description = '<br>'.join(i.description.split('\n')).replace(' ','&nbsp')
        aaData.append({
                       '0':objective,
                       '1':description,
                       '2':i.weight,
                       '3':i.self_report_value,
                       '4':i.self_report_score,
                       '5':i.supervisor_report_value,
                       '6':i.supervisor_report_score,
                       '7':i.principal_report_value,
                       '8':i.principal_report_score,
                       '9':i.KPI_name,
                       '10':i.name,
                       '11':i.id
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(simplejson.dumps(result),content_type="application/json")

@login_required
def KPI_table_detail_save(request):
    KPI_name = request.session.get('KPI')[0]
    name = request.session.get('KPI')[1]
    objective = request.POST.get('objective')
    description = request.POST.get ('description')
    weight = request.POST.get('weight')
    status = request.POST.get('status')
    _id = request.POST.get('id')

    ban_orm = ban.objects.filter(KPI_name=KPI_name)
    for i in ban_orm:
        if datetime.datetime.now().date() > i.ban_date:
            return HttpResponse(simplejson.dumps({'code':1,'msg':u'本季度KPI已经截止，无法保存'}),content_type="application/json")

    if not _id:
        orm = table_detail(KPI_name=KPI_name,name=name,objective=objective,description=description,weight=weight,
                           self_report_value=0,self_report_score=0,supervisor_report_value=0,supervisor_report_score=0,
                           principal_report_value=0,principal_report_score=0)
        try:
            orm.save()
            return HttpResponse(simplejson.dumps({'code':0,'msg':u'保存成功'}),content_type="application/json")
        except Exception,e:
            print e
            return HttpResponse(simplejson.dumps({'code':1,'msg':str(e)}),content_type="application/json")
    else:
        orm = table_detail.objects.get(id=_id)
        if int(status) == 3:
            self_report_value = int(request.POST.get('grade'))
            orm.self_report_value = self_report_value
            orm.self_report_score = float(self_report_value) / 100 * orm.weight
        if int(status) == 4:
            supervisor_report_value = int(request.POST.get('grade'))
            orm.supervisor_report_value = supervisor_report_value
            orm.supervisor_report_score = float(supervisor_report_value) / 100 * orm.weight
        if int(status) == 5:
            principal_report_value = int(request.POST.get('grade'))
            orm.principal_report_value = principal_report_value
            orm.principal_report_score = float(principal_report_value) / 100 * orm.weight
        orm.objective = objective
        orm.description = description
        orm.weight = weight
        try:
            orm.save()
            return HttpResponse(simplejson.dumps({'code':0,'msg':u'保存成功'}),content_type="application/json")
        except Exception,e:
            print e
            return HttpResponse(simplejson.dumps({'code':1,'msg':str(e)}),content_type="application/json")

@login_required
def KPI_table_detail_del(request):
    _id = request.POST.get('id')
    orm = table_detail.objects.get(id=_id)

    KPI_name = request.session.get('KPI')[0]
    ban_orm = ban.objects.filter(KPI_name=KPI_name)
    for i in ban_orm:
        if datetime.datetime.now().date() > i.ban_date:
            return HttpResponse(simplejson.dumps({'code':1,'msg':u'本季度KPI已经截止，无法保存'}),content_type="application/json")

    try:
        orm.delete()
        return HttpResponse(simplejson.dumps({'code':0,'msg':u'删除成功'}),content_type="application/json")
    except Exception,e:
        print e
        return HttpResponse(simplejson.dumps({'code':1,'msg':str(e)}),content_type="application/json")

@login_required
def KPI_table_detail_comment_save(request):
    self_comment = request.POST.get('self_comment')
    supervisor_comment = request.POST.get('supervisor_comment')
    principal_comment = request.POST.get('principal_comment')

    KPI_name = request.session.get('KPI')[0]
    name = request.session.get('KPI')[1]
    orm = table.objects.filter(KPI_name=KPI_name).filter(name=name)

    ban_orm = ban.objects.filter(KPI_name=KPI_name)
    for i in ban_orm:
        if datetime.datetime.now().date() > i.ban_date:
            return HttpResponse(simplejson.dumps({'code':1,'msg':u'本季度KPI已经截止，无法保存'}),content_type="application/json")

    if len(orm):
        for i in orm:
            if self_comment:
                i.self_comment = self_comment
            if supervisor_comment:
                i.supervisor_comment = supervisor_comment
            if principal_comment:
                i.principal_comment = principal_comment
            try:
                i.save()
                return HttpResponse(simplejson.dumps({'code':0,'msg':u'保存成功'}),content_type="application/json")
            except Exception,e:
                print e
                return HttpResponse(simplejson.dumps({'code':1,'msg':str(e)}),content_type="application/json")
    else:
        return HttpResponse(simplejson.dumps({'code':1,'msg':u'保存失败'}),content_type="application/json")

@login_required
def KPI_table_detail_commit(request):
    KPI_name = request.POST.get('KPI_name')
    name = request.POST.get('name')
    flag = request.POST.get('flag')

    ban_orm = ban.objects.filter(KPI_name=KPI_name)
    for i in ban_orm:
        if datetime.datetime.now().date() > i.ban_date:
            return HttpResponse(simplejson.dumps({'code':1,'msg':u'本季度KPI已经截止，无法保存'}),content_type="application/json")

    orm = table.objects.filter(KPI_name=KPI_name).filter(name=name)
    if len(orm):
        for i in orm:
            if flag == '0':
                i.status = 2
                vacation_user_table_orm = user_table.objects.get(name=name)
                i.commit_now = vacation_user_table_orm.supervisor
                i.status_interface = '等待 %s 确认目标' % vacation_user_table_orm.supervisor
                try:
                    vacation_user_table_orm = user_table.objects.get(name=i.commit_now)
                except Exception:
                    return HttpResponse(simplejson.dumps({'code':1,'msg':u'上级信息不存在'}),content_type="application/json")
                supervisor_email = vacation_user_table_orm.email
                vacation_user_table_orm.has_KPI_commit += 1
                if vacation_user_table_orm.KPI_commit_id:
                    if str(i.id) not in vacation_user_table_orm.KPI_commit_id.split(','):
                        vacation_user_table_orm.KPI_commit_id += ',' + str(i.id)
                else:
                    vacation_user_table_orm.KPI_commit_id = str(i.id)
                try:
                    i.save()
                    vacation_user_table_orm.save()
                    Thread(target=send_mail,args=(supervisor_email,'绩效审核提醒','<h3>有一个绩效事件等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
                    # send_mail(to_addr=supervisor_email,subject='绩效审核提醒',body='<h3>有一个绩效事件等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')
                    return HttpResponse(simplejson.dumps({'code':0,'msg':u'提交成功'}),content_type="application/json")
                except Exception,e:
                    print e
                    return HttpResponse(simplejson.dumps({'code':0,'msg':str(e)}),content_type="application/json")
            elif flag == '1':
                KPI_table_orm = table.objects.filter(name=name).filter(KPI_name=KPI_name)
                if len(KPI_table_orm) == 1:
                    for i in KPI_table_orm:
                        if i.self_comment == '':
                            return HttpResponse(simplejson.dumps({'code':1,'msg':u'请填写自我评价'}),content_type="application/json")
                vacation_user_table_orm = user_table.objects.get(name=name)
                if vacation_user_table_orm.supervisor == vacation_user_table_orm.principal:
                    i.status = 5
                else:
                    i.status = 4
                i.commit_now = vacation_user_table_orm.supervisor
                i.status_interface = '等待 %s 评分' % vacation_user_table_orm.supervisor

                try:
                    vacation_user_table_orm = user_table.objects.get(name=i.commit_now)
                except Exception:
                    return HttpResponse(simplejson.dumps({'code':1,'msg':u'上级信息不存在'}),content_type="application/json")
                supervisor_email = vacation_user_table_orm.email
                vacation_user_table_orm.has_KPI_commit += 1

                try:
                    i.save()
                    vacation_user_table_orm.save()
                    Thread(target=send_mail,args=(supervisor_email,'绩效审核提醒','<h3>有一个绩效事件等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
                    # send_mail(to_addr=supervisor_email,subject='绩效审核提醒',body='<h3>有一个绩效事件等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')
                    return HttpResponse(simplejson.dumps({'code':0,'msg':u'提交成功'}),content_type="application/json")
                except Exception,e:
                    print e
                    return HttpResponse(simplejson.dumps({'code':0,'msg':str(e)}),content_type="application/json")
            else:
                i.status = 7
                i.status_interface = '已完成'
                try:
                    i.save()
                    return HttpResponse(simplejson.dumps({'code':0,'msg':u'确认成功'}),content_type="application/json")
                except Exception,e:
                    print e
                    return HttpResponse(simplejson.dumps({'code':0,'msg':str(e)}),content_type="application/json")



@login_required
def KPI_approve_alert(request):
    orm = table.objects.filter(commit_now=request.user.first_name)
    if len(orm) > 0:
        msg = '有%s个绩效事件等待您的审核' % len(orm)
        return HttpResponse(simplejson.dumps({'code':0,'msg':msg}),content_type="application/json")
    else:
        return HttpResponse(simplejson.dumps({'code':1}),content_type="application/json")

@login_required
def KPI_table_approve(request):
    KPI_conf_save = request.GET.get('KPI_conf_save')
    path = request.path.split('/')[1]

    orm = ban.objects.all().order_by('id').reverse()
    if orm:
        expire_info = '''{0} 的截止日期为 {1}'''.format(orm[0].KPI_name,str(orm[0].ban_date).split('+')[0])
    else:
        expire_info = ''

    return render(request, 'KPI/KPI_table_approve.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'KPI',
                                                 'path2':path,
                                                 'page_name1':u'绩效管理',
                                                 'page_name2':u'绩效考评',
                                                 'username':request.user.username,
                                                 'KPI_conf_save':KPI_conf_save,
                                                 'expire_info':expire_info},
                                                context_instance=RequestContext(request))

@login_required
def KPI_table_approve_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['name','KPI_name','final_score','KPI_level','status_interface',None,'id']

    if request.user.has_perm('KPI.can_view_all'):
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                result_data = table.objects.all().order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().count()
            else:
                result_data = table.objects.all().filter(Q(name__contains=sSearch) | \
                                                        Q(KPI_name__contains=sSearch) | \
                                                        Q(final_score__contains=sSearch) | \
                                                        Q(KPI_level__contains=sSearch) | \
                                                        Q(status_interface__contains=sSearch)) \
                                                        .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().filter(Q(name__contains=sSearch) | \
                                                        Q(KPI_name__contains=sSearch) | \
                                                        Q(final_score__contains=sSearch) | \
                                                        Q(KPI_level__contains=sSearch) | \
                                                        Q(status_interface__contains=sSearch)).count()
        else:
            if sSearch == '':
                result_data = table.objects.all().order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().count()
            else:
                result_data = table.objects.all().filter(Q(name__contains=sSearch) | \
                                                        Q(KPI_name__contains=sSearch) | \
                                                        Q(final_score__contains=sSearch) | \
                                                        Q(KPI_level__contains=sSearch) | \
                                                        Q(status_interface__contains=sSearch)) \
                                                        .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().filter(Q(name__contains=sSearch) | \
                                                        Q(KPI_name__contains=sSearch) | \
                                                        Q(final_score__contains=sSearch) | \
                                                        Q(KPI_level__contains=sSearch) | \
                                                        Q(status_interface__contains=sSearch)).count()
    else:
        subordinate = []
        subordinate_orm = user_table.objects.filter(Q(supervisor=request.user.first_name) | Q(principal=request.user.first_name))
        for i in subordinate_orm:
            subordinate.append(i.name)

        orm_KPI_commit_id = user_table.objects.get(name=request.user.first_name)
        KPI_commit_id_list = orm_KPI_commit_id.KPI_commit_id.split(',')
        if KPI_commit_id_list != [u'']:
            KPI_commit_id_list = map(lambda x:int(x), KPI_commit_id_list)
        else:
            KPI_commit_id_list = []

        if  orm_KPI_commit_id.subordinate:
            subordinate_list = orm_KPI_commit_id.subordinate.split(',')
        else:
            subordinate_list = []

        if  sSortDir_0 == 'asc':
            if sSearch == '':
                result_data = table.objects.filter(Q(commit_now=request.user.first_name) | Q(name__in=subordinate)).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(Q(commit_now=request.user.first_name) | Q(name__in=subordinate)).count()
            else:
                result_data = table.objects.filter(Q(commit_now=request.user.first_name) | Q(name__in=subordinate)).filter(Q(KPI_name__contains=sSearch) | \
                                                        Q(final_score__contains=sSearch) | \
                                                        Q(status__contains=sSearch)) \
                                                        .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(Q(commit_now=request.user.first_name) | Q(name__in=subordinate)).filter(Q(KPI_name__contains=sSearch) | \
                                                        Q(final_score__contains=sSearch) | \
                                                        Q(status__contains=sSearch)).count()
        else:
            if sSearch == '':
                result_data = table.objects.filter(Q(commit_now=request.user.first_name) | Q(name__in=subordinate)).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(Q(commit_now=request.user.first_name) | Q(name__in=subordinate)).count()
            else:
                result_data = table.objects.filter(Q(commit_now=request.user.first_name) | Q(name__in=subordinate)).filter(Q(KPI_name__contains=sSearch) | \
                                                        Q(final_score__contains=sSearch) | \
                                                        Q(status__contains=sSearch)) \
                                                        .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(Q(commit_now=request.user.first_name) | Q(name__in=subordinate)).filter(Q(KPI_name__contains=sSearch) | \
                                                        Q(final_score__contains=sSearch) | \
                                                        Q(status__contains=sSearch)).count()



    for i in  result_data:
        export = '''
                <a class="btn btn-sm green"">
                    生成Excel文件 <i class="fa fa-level-down"></i>
                </a>
            '''
        aaData.append({
                       '0':i.name,
                       '1':i.KPI_name,
                       '2':i.final_score,
                       '3':i.KPI_level,
                       '4':i.status_interface,
                       '5':export,
                       '6':i.status,
                       '7':i.id
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(simplejson.dumps(result),content_type="application/json")

@login_required
def KPI_table_detail_approve(request):
    KPI_name = request.session.get('KPI')[0]
    name = request.session.get('KPI')[1]
    orm = table.objects.filter(KPI_name=KPI_name).filter(name=name)

    if len(orm):
        for i in orm:
            self_comment = i.self_comment.replace('\n','\\n')
            supervisor_comment = i.supervisor_comment.replace('\n','\\n')
            principal_comment = i.principal_comment.replace('\n','\\n')
            commit_now = i.commit_now
    else:
        self_comment = ''
        supervisor_comment = ''
        principal_comment = ''
        commit_now = ""
    # try:
    #     if status:pass
    # except Exception:
    #     status = '员工设定目标'
    return render(request, 'KPI/KPI_table_detail_approve.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'KPI',
                                                 'path2':'',
                                                 'page_name1':u'绩效管理',
                                                 'page_name2':u'绩效考评详情',
                                                 'self_comment':self_comment,
                                                 'supervisor_comment':supervisor_comment,
                                                 'principal_comment':principal_comment,
                                                 'KPI_name':KPI_name,
                                                 'name':name,
                                                 'commit_now':commit_now,
                                                 'self_comment_row':len(self_comment.split('\\n')) + 2,
                                                 'supervisor_comment_row':len(supervisor_comment.split('\\n')) + 2,
                                                 'principal_comment_row':len(principal_comment.split('\\n')) + 2},
                                                context_instance=RequestContext(request))

@login_required
def KPI_table_detail_approve_commit(request):
    KPI_name = request.POST.get('KPI_name')
    name = request.POST.get('name')
    flag = request.POST.get('flag')
    commit = request.POST.get('commit')

    ban_orm = ban.objects.filter(KPI_name=KPI_name)
    for i in ban_orm:
        if datetime.datetime.now().date() > i.ban_date:
            return HttpResponse(simplejson.dumps({'code':1,'msg':u'本季度KPI已经截止，无法保存'}),content_type="application/json")

    orm = table.objects.filter(KPI_name=KPI_name).filter(name=name)
    if len(orm):
        for i in orm:
            if not int(flag):
                if int(commit):
                    i.status = 3
                    i.commit_now = ''
                    i.status_interface = '员工自我评分'

                    vacation_user_table_orm = user_table.objects.get(name=name)
                    email = vacation_user_table_orm.email

                    vacation_user_table_orm2 = user_table.objects.get(name=request.user.first_name)
                    vacation_user_table_orm2.has_KPI_commit -= 1
                    try:
                        i.save()
                        vacation_user_table_orm2.save()
                        Thread(target=send_mail,args=(email,'绩效审核提醒','<h3>%s 通过了您设定的绩效目标，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。' % request.user.first_name)).start()
                        # send_mail(to_addr=email,subject='绩效审核提醒',body='<h3>%s 通过了您设定的绩效目标，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。' % request.user.first_name)
                        return HttpResponse(simplejson.dumps({'code':0,'msg':u'提交成功'}),content_type="application/json")
                    except Exception,e:
                        print e
                        return HttpResponse(simplejson.dumps({'code':0,'msg':str(e)}),content_type="application/json")
                else:
                    i.status = 1
                    i.commit_now = ''
                    i.status_interface = '员工设定目标'

                    vacation_user_table_orm = user_table.objects.get(name=name)
                    email = vacation_user_table_orm.email

                    vacation_user_table_orm2 = user_table.objects.get(name=request.user.first_name)
                    vacation_user_table_orm2.has_KPI_commit -= 1

                    reason = request.POST.get('reason')
                    try:
                        i.save()
                        vacation_user_table_orm2.save()
                        Thread(target=send_mail,args=(email,'绩效审核提醒','<h3>%s 不通过您设定的绩效目标，请在OA系统中查看。</h3><br>拒绝理由：<font color="red">%s</font><br><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。' % (request.user.first_name,reason))).start()
                        # send_mail(to_addr=email,subject='绩效审核提醒',body='<h3>%s 不通过您设定的绩效目标，请在OA系统中查看。</h3><br>拒绝理由：<font color="red">%s</font><br><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。' % (request.user.first_name,reason))
                        return HttpResponse(simplejson.dumps({'code':0,'msg':u'提交成功'}),content_type="application/json")
                    except Exception,e:
                        print e
                        return HttpResponse(simplejson.dumps({'code':0,'msg':str(e)}),content_type="application/json")
            else:
                if i.status == 4:
                    i.status = 5
                    vacation_user_table_orm = user_table.objects.get(name=name)
                    i.commit_now = vacation_user_table_orm.principal
                    i.status_interface = '等待 %s 评分' % vacation_user_table_orm.principal

                    vacation_user_table_orm2 = user_table.objects.get(name=i.commit_now)
                    principal_email = vacation_user_table_orm2.email
                    vacation_user_table_orm2.has_KPI_commit += 1

                    try:
                        i.save()
                        vacation_user_table_orm2.save()
                        vacation_user_table_orm3 = user_table.objects.get(name=request.user.first_name)
                        vacation_user_table_orm3.has_KPI_commit -= 1
                        vacation_user_table_orm3.save()
                        Thread(target=send_mail,args=(principal_email,'绩效审核提醒','<h3>有一个绩效事件等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
                        # send_mail(to_addr=principal_email,subject='绩效审核提醒',body='<h3>有一个绩效事件等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')
                        return HttpResponse(simplejson.dumps({'code':0,'msg':u'提交成功'}),content_type="application/json")
                    except Exception,e:
                        print e
                        return HttpResponse(simplejson.dumps({'code':0,'msg':str(e)}),content_type="application/json")
                if i.status == 5:
                    if i.commit_now == request.user.first_name:
                        i.status = 6
                        i.commit_now = ''
                        i.status_interface = '等待员工最终确认'

                        vacation_user_table_orm = user_table.objects.get(name=name)
                        email = vacation_user_table_orm.email

                        vacation_user_table_orm2 = user_table.objects.get(name=request.user.first_name)
                        vacation_user_table_orm2.has_KPI_commit -= 1

                        final_score = request.POST.get('sum')
                        i.final_score = float(final_score)

                        try:
                            i.save()
                            vacation_user_table_orm2.save()
                            Thread(target=send_mail,args=(email,'绩效审核提醒','<h3>您的绩效考评已完成全部打分，等待最终确认，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
                            # send_mail(to_addr=email,subject='绩效审核提醒',body='<h3>您的绩效考评已完成全部打分，等待最终确认，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')
                            return HttpResponse(simplejson.dumps({'code':0,'msg':u'提交成功'}),content_type="application/json")
                        except Exception,e:
                            print e
                            return HttpResponse(simplejson.dumps({'code':0,'msg':str(e)}),content_type="application/json")
                    else:
                        return HttpResponse(simplejson.dumps({'code':0,'msg':'当前审核人不是您'}),content_type="application/json")

@login_required
def create_excel(requests):
    _id = requests.POST.get('id')
    KPI_table_orm = table.objects.get(id=_id)
    user_table_orm = user_table.objects.get(name=KPI_table_orm.name)
    KPI_table_detail_orm_iter = table_detail.objects.filter(name=KPI_table_orm.name,KPI_name=KPI_table_orm.KPI_name)

    if KPI_table_orm.status != 7:
        return HttpResponse(simplejson.dumps({'code':1,'msg':'您的绩效未完成无法生成Excel'}),content_type="application/json")

    name = user_table_orm.name
    join_date = user_table_orm.join_date
    supervisor = user_table_orm.supervisor
    KPI_name = KPI_table_orm.KPI_name
    final_score = KPI_table_orm.final_score
    KPI_level = KPI_table_orm.KPI_level
    self_comment = KPI_table_orm.self_comment
    supervisor_comment = KPI_table_orm.supervisor_comment
    principal_comment = KPI_table_orm.principal_comment

    objective_list = []
    for KPI_detail in KPI_table_detail_orm_iter:
        objective_list.append({'objective':KPI_detail.objective,'description':KPI_detail.description,
                               'weight':KPI_detail.weight,'self_report_value':KPI_detail.self_report_value,
                               'supervisor_report_value':KPI_detail.supervisor_report_value,
                               'principal_report_value':KPI_detail.principal_report_value,
                               'self_report_score':KPI_detail.self_report_score,
                               'supervisor_report_score':KPI_detail.supervisor_report_score,
                               'principal_report_score':KPI_detail.principal_report_score})

    try:
        save_dir = BASE_DIR + '/static/files/KPI/' + requests.user.username
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        wb = load_workbook(filename = BASE_DIR + '/static/files/KPI_template.xlsx')
        ws = wb.active
        ws['B6'] = name
        ws['F6'] = join_date
        ws['B7'] = supervisor
        ws['F7'] = KPI_name

        row_num = '10'
        self_total_score = 0
        supervisor_total_score = 0
        principal_total_score = 0
        for objective in objective_list:
            ws['B'+row_num] = objective['objective'].replace('<br>','\n').replace('&nbsp',' ')
            ws['C'+row_num] = objective['description'].replace('<br>','\n').replace('&nbsp',' ')
            ws['E'+row_num] = str(objective['weight']) + '%'
            ws['F'+row_num] = objective['self_report_value']
            ws['G'+row_num] = objective['supervisor_report_value']
            ws['H'+row_num] = objective['principal_report_value']
            row_num = str(int(row_num) + 2)
            self_total_score += objective['self_report_score']
            supervisor_total_score += objective['supervisor_report_score']
            principal_total_score += objective['principal_report_score']

        ws['F20'] = self_total_score
        ws['G20'] = supervisor_total_score
        ws['H20'] = principal_total_score
        ws['H21'] = final_score
        ws['H22'] = KPI_level
        ws['B23'] = self_comment
        ws['B24'] = supervisor_comment
        ws['B25'] = principal_comment

        wb.save(save_dir + '/KPI.xlsx')
        return HttpResponse(simplejson.dumps({'code':0,'msg':u'生成成功'}),content_type="application/json")
    except Exception,e:
        print e
        return HttpResponse(simplejson.dumps({'code':1,'msg':str(e)}),content_type="application/json")

@login_required
def KPI_upload_conf(requests):
    KPI_conf = requests.FILES.get('KPI_conf')
    try:
        with open(BASE_DIR+'/static/files/KPI_conf','w') as f:
            for data in KPI_conf.chunks():
                f.write(data.decode('gbk').encode('utf8'))

        with open(BASE_DIR+'/static/files/KPI_conf') as f:
            line_1_list = f.readline().split()
            KPI_name = line_1_list[0]
            print line_1_list[1]
            if line_1_list[1] == "绩效等级":
                line = f.readline()
                while line:
                    if line == '':
                        continue
                    line_list = line.split()
                    orm = table.objects.filter(name=line_list[0]).filter(KPI_name=KPI_name)
                    if len(orm) == 1:
                        for i in orm:
                            i.KPI_level = line_list[1]
                            i.save()
                    line = f.readline()
            elif line_1_list[1] == "员工信息":
                line = f.readline()
                while line:
                    if line == '':
                        continue
                    name = line.strip()
                    if len(table.objects.filter(KPI_name=KPI_name).filter(name=name)) == 0:
                        orm = table(KPI_name=KPI_name,name=name,final_score=0,status_interface='员工设定目标',status=1)
                        orm.save()
                    line = f.readline()
        return HttpResponseRedirect('/KPI_table_approve/?KPI_conf_save=1')
    except Exception,e:
        print e
        return HttpResponseRedirect('/KPI_table_approve/?KPI_conf_save=2')

@login_required
def KPI_ban_save(requests):
    ban_KPI_name = requests.POST.get('ban_KPI_name')
    ban_date = requests.POST.get('ban_date')
    ban_type = requests.POST.get('ban_type')

    if ban_type == '1':
        if not ban_date:
            ban_date = datetime.datetime.now().date()
        orm = ban(KPI_name=ban_KPI_name,ban_date=ban_date)
        try:
            orm.save()
            return HttpResponse(simplejson.dumps({'code':0,'msg':u'封禁成功'}),content_type="application/json")
        except Exception,e:
            print e
            return HttpResponse(simplejson.dumps({'code':1,'msg':str(e)}),content_type="application/json")
    else:
        if ban_date:
            orm = ban.objects.filter(KPI_name=ban_KPI_name).filter(ban_date=ban_date)
        else:
            orm = ban.objects.filter(KPI_name=ban_KPI_name)
        if not len(orm):
            return HttpResponse(simplejson.dumps({'code':1,'msg':'没有可解封的对象'}),content_type="application/json")
        else:
            try:
                for i in orm:
                    i.delete()
                return HttpResponse(simplejson.dumps({'code':0,'msg':u'解封成功'}),content_type="application/json")
            except Exception,e:
                print e
                return HttpResponse(simplejson.dumps({'code':1,'msg':str(e)}),content_type="application/json")
